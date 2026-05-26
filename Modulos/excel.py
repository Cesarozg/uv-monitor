import openpyxl
import pandas as pd
def crear_tabla_excel(datos_ciudades):
  df = pd.DataFrame(datos_ciudades)
  
  # Crear un escritor de pandas Excel con formato
  with pd.ExcelWriter('datos_ciudades.xlsx', engine='openpyxl') as writer:
      df.to_excel(writer, index=False, sheet_name='Datos Ciudades')
  
      # Obtener la hoja de trabajo
      sheet = writer.sheets['Datos Ciudades']
  
      # Establecer el formato condicional para resaltar los valores más altos
      max_uv = df['uv'].max()
      max_radiacion = df['radiacion'].max()
      max_ozono = df['ozono'].max()
      for idx, row in df.iterrows():
          if row['uv'] == max_uv:
              sheet.cell(row=idx + 2, column=2).font = openpyxl.styles.Font(bold=True, color="FF0000")
          if row['radiacion'] == max_radiacion:
              sheet.cell(row=idx + 2, column=3).font = openpyxl.styles.Font(bold=True, color="FF0000")
          if row['radiacion'] is not None and row['radiacion'] > 8:
              sheet.cell(row=idx + 2, column=3).fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
          if row['ozono'] == max_ozono:
            sheet.cell(row=idx + 2, column=3).font = openpyxl.styles.Font(bold=True, color="FF0000")
          if row['ozono'] is not None and row['ozono'] > 8:
            sheet.cell(row=idx + 2, column=3).fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
  
      # Centrar el contenido en las celdas
      for column in sheet.columns:
          max_length = 0
          for cell in column:
              try:
                  if len(str(cell.value)) > max_length:
                      max_length = len(cell.value)
              except:
                  pass
          adjusted_width = (max_length + 2)
          sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
  
  print("Tabla Excel creada con éxito: datos_ciudades.xlsx")